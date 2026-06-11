from pathlib import Path
import json
import re
import subprocess
import sys

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
FINAL_DIR = OUTPUTS / "战术事件_序号1-5_v1.9"
DEFINITION_PATH = OUTPUTS / "dota2_tactical_event_definitions_v1.9_20260609.xlsx"
MATCH_IDS = ["8825993964", "8825996999", "8826042346", "8826052816", "8826099852"]
EXPECTED_HEADERS = ["id", "match_id", "labels", "confidence", "time_range", "heroes", "结果", "批注"]


def split_labels(value):
    return [label.strip() for label in re.split(r"\s+/\s+", str(value or "")) if label.strip()]


def has_border(cell):
    return all(getattr(cell.border, side).style for side in ("left", "right", "top", "bottom"))


def workbook_texts(workbook):
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    yield str(cell.value)


def load_allowed_labels():
    wb = load_workbook(DEFINITION_PATH, read_only=True, data_only=False)
    texts = list(workbook_texts(wb))
    if any("抓人" in text for text in texts):
        raise AssertionError("v1.9 definition still contains 抓人")
    labels = set()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        first = next(rows, None)
        if first and first[0] in ("label", "标签"):
            for row in rows:
                if row and row[0]:
                    labels.add(str(row[0]))
    if len(labels) != 18 or "肉山击杀" not in labels:
        raise AssertionError(f"Expected 18 labels including 肉山击杀, got {len(labels)}")
    return labels


def verify_event_json(match_id, allowed_labels):
    path = OUTPUTS / f"tactical_events_{match_id}" / f"tactical_events_{match_id}_event_table_clean_v1.9.json"
    rows = json.loads(path.read_text(encoding="utf-8"))
    for row in rows:
        labels = split_labels(row["labels"])
        if any("抓人" in label for label in labels):
            raise AssertionError(f"{match_id} JSON contains 抓人")
        missing = set(labels) - allowed_labels
        if missing:
            raise AssertionError(f"{match_id} JSON has labels outside v1.9 definitions: {sorted(missing)}")
        result = str(row.get("结果", ""))
        if "资源获得" in labels and ("击杀肉山" in result or "获得肉山盾" in result or "获得奶酪" in result or "获得刷新碎片" in result):
            raise AssertionError(f"{match_id} Roshan result still under 资源获得: {row}")
        if "肉山击杀" in labels and ("击杀肉山" not in result or "获得" not in result):
            raise AssertionError(f"{match_id} 肉山击杀 result incomplete: {result}")
        if "批注" not in row:
            raise AssertionError(f"{match_id} JSON missing 批注 field")
    return len(rows)


def verify_workbook(match_id, expected_rows):
    path = FINAL_DIR / f"战术事件_{match_id}.xlsx"
    if not path.exists():
        raise AssertionError(f"Missing final workbook: {path}")
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb["战术事件"]
    headers = [ws.cell(1, col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    if headers != EXPECTED_HEADERS:
        raise AssertionError(f"{match_id} header mismatch: {headers}")
    if ws.max_row - 1 != expected_rows:
        raise AssertionError(f"{match_id} row mismatch: xlsx={ws.max_row - 1}, json={expected_rows}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(EXPECTED_HEADERS)):
        for cell in row:
            if not has_border(cell):
                raise AssertionError(f"{match_id} missing border at {cell.coordinate}")
    labels_col = EXPECTED_HEADERS.index("labels") + 1
    comment_col = EXPECTED_HEADERS.index("批注") + 1
    for row in range(2, ws.max_row + 1):
        if "抓人" in str(ws.cell(row, labels_col).value or ""):
            raise AssertionError(f"{match_id} workbook contains 抓人 at row {row}")
        if ws.cell(row, comment_col).value not in (None, ""):
            raise AssertionError(f"{match_id} 批注 column is not blank at row {row}")
    meta = wb["定义版本"]
    meta_rows = {(str(meta.cell(row, 1).value), str(meta.cell(row, 2).value)) for row in range(1, meta.max_row + 1)}
    if ("定义版本", "v1.9") not in meta_rows:
        raise AssertionError(f"{match_id} workbook metadata is not v1.9")
    return {"match_id": match_id, "rows": expected_rows, "file": path.name}


def verify_adjacent_diagnostic_empty():
    proc = subprocess.run(
        [sys.executable, str(ROOT / "work" / "diagnose_adjacent_combats_v1_9.py"), "v1.9"],
        cwd=str(ROOT),
        text=True,
        capture_output=True,
        check=True,
    )
    data = json.loads(proc.stdout)
    leftovers = {match_id: rows for match_id, rows in data.items() if rows}
    if leftovers:
        raise AssertionError(f"Adjacent combat diagnostic still has candidates: {json.dumps(leftovers, ensure_ascii=False)}")


def main():
    files = sorted(path.name for path in FINAL_DIR.glob("*.xlsx"))
    expected_files = [f"战术事件_{match_id}.xlsx" for match_id in MATCH_IDS]
    if files != expected_files:
        raise AssertionError(f"Final folder file mismatch: {files}")
    allowed_labels = load_allowed_labels()
    summaries = []
    for match_id in MATCH_IDS:
        row_count = verify_event_json(match_id, allowed_labels)
        summaries.append(verify_workbook(match_id, row_count))
    verify_adjacent_diagnostic_empty()
    print(json.dumps({
        "definition": str(DEFINITION_PATH),
        "definition_labels": len(allowed_labels),
        "final_dir": str(FINAL_DIR),
        "files": files,
        "matches": summaries,
        "adjacent_combat_candidates": 0,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
