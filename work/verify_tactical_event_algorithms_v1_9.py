from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PATH = ROOT / "outputs" / "dota2_tactical_event_recognition_algorithms_v1.9_20260609.xlsx"
EXPECTED_SHEETS = [
    "版本说明",
    "标签算法",
    "参数阈值",
    "数据源字段",
    "处理流程",
    "合并去重",
    "输出格式",
    "人工修改入口",
]


def main():
    workbook = load_workbook(PATH)
    label_sheet = workbook["标签算法"]
    labels = [
        label_sheet.cell(row=row, column=1).value
        for row in range(2, label_sheet.max_row + 1)
        if label_sheet.cell(row=row, column=1).value
    ]

    missing_border = []
    for sheet_name in EXPECTED_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                border = cell.border
                if not (border.left.style and border.right.style and border.top.style and border.bottom.style):
                    missing_border.append(f"{sheet_name}!{cell.coordinate}")
                    break
            if missing_border:
                break

    print({
        "path": str(PATH),
        "exists": PATH.exists(),
        "size": PATH.stat().st_size if PATH.exists() else 0,
        "sheets": workbook.sheetnames,
        "expected_sheets": workbook.sheetnames == EXPECTED_SHEETS,
        "label_count": len(labels),
        "has_gank_label": "抓人" in labels,
        "has_roshan_kill": "肉山击杀" in labels,
        "parameter_rows": workbook["参数阈值"].max_row - 1,
        "merge_rows": workbook["合并去重"].max_row - 1,
        "output_rows": workbook["输出格式"].max_row - 1,
        "missing_border_count": len(missing_border),
        "missing_border_sample": missing_border[:5],
    })

    assert workbook.sheetnames == EXPECTED_SHEETS
    assert len(labels) == 18
    assert "抓人" not in labels
    assert "肉山击杀" in labels
    assert workbook["参数阈值"].max_row - 1 >= 60
    assert workbook["合并去重"].max_row - 1 >= 10
    assert not missing_border


if __name__ == "__main__":
    main()
